"""Pilot Routes: Pilot Personal Records, Training Syllabus, Training Status, Weekly Report"""
import os
import io
import json
import re
import time
import uuid
import datetime
from database import get_db, dict_from_row, dicts_from_rows
from auth import require_auth, get_current_user
from routes.auth_routes import BaseHandler

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "public", "uploads")
WEEKLY_UPLOAD_DIR = os.path.join(UPLOAD_DIR, "weekly")

PILOT_FIELDS = [
    'name', 'short_name', 'rank', 'service_number', 'callsign', 'nationality',
    'squadron', 'date_of_birth', 'course_class', 'training_start_date',
    'training_end_date', 'phone', 'email', 'photo_url', 'notes', 'sort_order', 'status',
]


class PilotsHandler(BaseHandler):
    """GET list, POST create"""

    @require_auth()
    def get(self):
        conn = get_db()
        status = self.get_argument("status", None)
        if status:
            pilots = dicts_from_rows(conn.execute(
                "SELECT * FROM pilots WHERE status=? ORDER BY sort_order, id", (status,)
            ).fetchall())
        else:
            pilots = dicts_from_rows(conn.execute(
                "SELECT * FROM pilots ORDER BY sort_order, id"
            ).fetchall())
        conn.close()
        self.success(pilots)

    @require_auth(roles=['admin', 'ojt_admin'])
    def post(self):
        body = self.get_json_body()
        name = body.get('name', '').strip()
        short_name = body.get('short_name', '').strip()
        if not name or not short_name:
            return self.error("Name and short name are required")

        conn = get_db()
        cur = conn.execute(
            """INSERT INTO pilots (name, short_name, rank, service_number, callsign,
               nationality, squadron, date_of_birth, course_class,
               training_start_date, training_end_date, phone, email, notes, sort_order)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (name, short_name, body.get('rank', 'Major'),
             body.get('service_number', ''), body.get('callsign', ''),
             body.get('nationality', 'Malaysia'), body.get('squadron', ''),
             body.get('date_of_birth', ''), body.get('course_class', ''),
             body.get('training_start_date', ''), body.get('training_end_date', ''),
             body.get('phone', ''), body.get('email', ''),
             body.get('notes', ''), body.get('sort_order', 0))
        )
        conn.commit()
        pilot = dict_from_row(conn.execute("SELECT * FROM pilots WHERE id=?", (cur.lastrowid,)).fetchone())
        conn.close()
        self.success(pilot, "Pilot created")


class PilotDetailHandler(BaseHandler):
    """GET one, PUT update, DELETE deactivate"""

    @require_auth()
    def get(self, pilot_id):
        conn = get_db()
        pilot = dict_from_row(conn.execute("SELECT * FROM pilots WHERE id=?", (pilot_id,)).fetchone())
        conn.close()
        if not pilot:
            return self.error("Pilot not found", 404)
        self.success(pilot)

    @require_auth(roles=['admin', 'ojt_admin'])
    def put(self, pilot_id):
        body = self.get_json_body()
        conn = get_db()
        existing = conn.execute("SELECT id FROM pilots WHERE id=?", (pilot_id,)).fetchone()
        if not existing:
            conn.close()
            return self.error("Pilot not found", 404)

        fields, vals = [], []
        for k in PILOT_FIELDS:
            if k in body:
                fields.append(f"{k}=?")
                vals.append(body[k])
        if not fields:
            conn.close()
            return self.error("No fields to update")

        fields.append("updated_at=CURRENT_TIMESTAMP")
        vals.append(pilot_id)
        conn.execute(f"UPDATE pilots SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
        pilot = dict_from_row(conn.execute("SELECT * FROM pilots WHERE id=?", (pilot_id,)).fetchone())
        conn.close()
        self.success(pilot, "Pilot updated")

    @require_auth(roles=['admin', 'ojt_admin'])
    def delete(self, pilot_id):
        conn = get_db()
        existing = conn.execute("SELECT id, name FROM pilots WHERE id=?", (pilot_id,)).fetchone()
        if not existing:
            conn.close()
            return self.error("Pilot not found", 404)
        conn.execute("UPDATE pilots SET status='inactive', updated_at=CURRENT_TIMESTAMP WHERE id=?", (pilot_id,))
        conn.commit()
        conn.close()
        self.success(message="Pilot deactivated")


class PilotPhotoHandler(BaseHandler):
    """POST upload pilot photo"""

    @require_auth(roles=['admin', 'ojt_admin'])
    def post(self, pilot_id):
        conn = get_db()
        pilot = conn.execute("SELECT id FROM pilots WHERE id=?", (pilot_id,)).fetchone()
        if not pilot:
            conn.close()
            return self.error("Pilot not found", 404)

        files = self.request.files.get("photo", [])
        if not files:
            conn.close()
            return self.error("No photo file uploaded")

        f = files[0]
        ext = os.path.splitext(f["filename"])[1].lower()
        if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
            conn.close()
            return self.error("Unsupported file type")
        if len(f["body"]) > 5 * 1024 * 1024:
            conn.close()
            return self.error("File too large (max 5MB)")

        os.makedirs(UPLOAD_DIR, exist_ok=True)
        fname = f"pilot_{pilot_id}_{uuid.uuid4().hex[:8]}{ext}"
        fpath = os.path.join(UPLOAD_DIR, fname)
        with open(fpath, "wb") as fp:
            fp.write(f["body"])

        photo_url = f"/uploads/{fname}"
        conn.execute("UPDATE pilots SET photo_url=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (photo_url, pilot_id))
        conn.commit()
        conn.close()
        self.success({"photo_url": photo_url}, "Photo uploaded")


class PilotCoursesHandler(BaseHandler):
    """GET all training syllabus courses"""

    @require_auth()
    def get(self):
        conn = get_db()
        courses = dicts_from_rows(conn.execute(
            "SELECT * FROM pilot_courses ORDER BY sort_order, id"
        ).fetchall())
        conn.close()
        self.success(courses)


class PilotTrainingHandler(BaseHandler):
    """GET all training records, POST upsert a record"""

    @require_auth()
    def get(self):
        conn = get_db()
        records = dicts_from_rows(conn.execute("""
            SELECT pt.id, pt.pilot_id, pt.course_id, pt.completed_date,
                   pt.completed_time, pt.notes, pt.updated_at,
                   p.short_name as pilot_name,
                   pc.subject, pc.category, pc.course_no
            FROM pilot_training pt
            JOIN pilots p ON pt.pilot_id = p.id
            JOIN pilot_courses pc ON pt.course_id = pc.id
            ORDER BY p.sort_order, pc.sort_order
        """).fetchall())
        conn.close()
        self.success(records)

    @require_auth(roles=['admin', 'ojt_admin', 'instructor'])
    def post(self):
        body = self.get_json_body()
        pilot_id = body.get('pilot_id')
        course_id = body.get('course_id')
        completed_date = body.get('completed_date')
        completed_time = body.get('completed_time', '1:00')
        notes = body.get('notes', '')

        if not pilot_id or not course_id:
            return self.error("pilot_id and course_id required")

        conn = get_db()
        existing = conn.execute(
            "SELECT id FROM pilot_training WHERE pilot_id=? AND course_id=?",
            (pilot_id, course_id)
        ).fetchone()

        if existing:
            if completed_date:
                conn.execute(
                    """UPDATE pilot_training SET completed_date=?, completed_time=?,
                       notes=?, updated_at=CURRENT_TIMESTAMP
                       WHERE pilot_id=? AND course_id=?""",
                    (completed_date, completed_time, notes, pilot_id, course_id)
                )
            else:
                conn.execute(
                    "DELETE FROM pilot_training WHERE pilot_id=? AND course_id=?",
                    (pilot_id, course_id)
                )
        elif completed_date:
            conn.execute(
                """INSERT INTO pilot_training
                   (pilot_id, course_id, completed_date, completed_time, notes)
                   VALUES (?,?,?,?,?)""",
                (pilot_id, course_id, completed_date, completed_time, notes)
            )
        conn.commit()
        conn.close()
        self.success(message="Training record updated")


class PilotWeeklyHandler(BaseHandler):
    """GET weekly summary per pilot — uses latest upload if available, otherwise computed from training records"""

    @require_auth()
    def get(self):
        conn = get_db()

        # Check if there's uploaded weekly data
        latest_upload = conn.execute(
            "SELECT id FROM weekly_uploads ORDER BY created_at DESC LIMIT 1"
        ).fetchone()

        if latest_upload:
            # Use data from latest upload
            upload_data = dicts_from_rows(conn.execute(
                "SELECT * FROM weekly_report_data WHERE upload_id=? ORDER BY id",
                (latest_upload['id'],)
            ).fetchall())
            result = []
            for d in upload_data:
                result.append({
                    'id': d['pilot_id'] or 0,
                    'name': d['pilot_name'],
                    'simPlan': d['sim_plan'], 'simDone': d['sim_done'], 'simRemain': d['sim_remain'],
                    'fltPlan': d['flt_plan'], 'fltDone': d['flt_done'], 'fltRemain': d['flt_remain'],
                })
            conn.close()
            return self.success(result)

        # Fallback: compute from training records
        pilots = dicts_from_rows(conn.execute(
            "SELECT * FROM pilots WHERE status='active' ORDER BY sort_order, id"
        ).fetchall())

        sim_total = conn.execute("SELECT COUNT(*) as cnt FROM pilot_courses WHERE category='sim'").fetchone()['cnt']
        flt_total = conn.execute("SELECT COUNT(*) as cnt FROM pilot_courses WHERE category='flight'").fetchone()['cnt']

        result = []
        for p in pilots:
            sim_done = conn.execute("""
                SELECT COUNT(*) as cnt FROM pilot_training pt
                JOIN pilot_courses pc ON pt.course_id = pc.id
                WHERE pt.pilot_id=? AND pc.category='sim' AND pt.completed_date IS NOT NULL
            """, (p['id'],)).fetchone()['cnt']
            flt_done = conn.execute("""
                SELECT COUNT(*) as cnt FROM pilot_training pt
                JOIN pilot_courses pc ON pt.course_id = pc.id
                WHERE pt.pilot_id=? AND pc.category='flight' AND pt.completed_date IS NOT NULL
            """, (p['id'],)).fetchone()['cnt']
            result.append({
                'id': p['id'], 'name': p['short_name'],
                'simPlan': sim_total, 'simDone': sim_done, 'simRemain': sim_total - sim_done,
                'fltPlan': flt_total, 'fltDone': flt_done, 'fltRemain': flt_total - flt_done,
            })
        conn.close()
        self.success(result)


class PilotNationalitiesHandler(BaseHandler):
    """GET list of nationalities, POST add new nationality"""

    @require_auth()
    def get(self):
        conn = get_db()
        rows = dicts_from_rows(conn.execute(
            "SELECT * FROM pilot_nationalities ORDER BY sort_order, id"
        ).fetchall())
        conn.close()
        self.success(rows)

    @require_auth()
    def post(self):
        body = self.get_json_body()
        code = (body.get('code') or '').strip()
        label_ko = (body.get('label_ko') or '').strip()
        if not code or not label_ko:
            return self.error("code and label_ko are required")
        conn = get_db()
        existing = conn.execute("SELECT id FROM pilot_nationalities WHERE code=?", (code,)).fetchone()
        if existing:
            conn.close()
            return self.error("Nationality already exists")
        sort_order = body.get('sort_order', 0)
        cur = conn.execute(
            "INSERT INTO pilot_nationalities (code, label_ko, sort_order) VALUES (?,?,?)",
            (code, label_ko, sort_order)
        )
        conn.commit()
        row = dict_from_row(conn.execute("SELECT * FROM pilot_nationalities WHERE id=?", (cur.lastrowid,)).fetchone())
        conn.close()
        self.success(row, "Nationality added")


class WeeklyUploadHandler(BaseHandler):
    """GET list uploads, POST upload new Excel file"""

    def _safe_int(self, val):
        try:
            return int(float(val)) if val is not None else 0
        except (ValueError, TypeError):
            return 0

    def _parse_weekly_report_sheet(self, ws):
        """Parse RMAF-style horizontal weekly report.
        Layout:
          Row 5: pilot names at columns H,K,N,Q,T,W,... (3-col stride: Plan/Done/Remain)
          Row 7: header row with Plan/Done/Remain repeated
          Row 8: Flight Sortie data
          Row 10: Simulator Sortie data
        """
        parsed = []
        # Step 1: Find pilot name row — scan rows 3-8 for cells with long text names
        name_row = None
        pilot_cols = []
        for r in range(3, 10):
            row_cells = list(ws.iter_rows(min_row=r, max_row=r, values_only=False))[0]
            candidates = []
            for cell in row_cells:
                v = cell.value
                if v and isinstance(v, str) and len(v.strip()) > 3 and not any(
                    kw in v.lower() for kw in ('week', 'result', 'item', 'plan', 'done', 'remain', 'report', 'total',
                                                  'flight', 'sim', 'sum', 'sortie', 'time', 'course', 'as of',
                                                  '주간', '소계', '합계', '비행', '학술')
                ) and cell.column >= 5:
                    candidates.append(cell)
            if len(candidates) >= 2:
                name_row = r
                pilot_cols = candidates
                break

        if not name_row or len(pilot_cols) < 2:
            return []

        # Step 2: Find data rows — Flight Sortie and Simulator Sortie
        flt_row = None
        sim_row = None
        for r in range(name_row + 1, min(name_row + 15, ws.max_row + 1)):
            for cell in ws.iter_rows(min_row=r, max_row=r, values_only=False):
                for c in cell:
                    v = str(c.value or '').strip().lower()
                    if v == 'flight' and flt_row is None:
                        # Check next column for 'sortie'
                        next_cell = ws.cell(row=r, column=c.column + 1)
                        nv = str(next_cell.value or '').strip().lower()
                        if nv == 'sortie':
                            flt_row = r
                    elif v == 'simulator' and sim_row is None:
                        next_cell = ws.cell(row=r, column=c.column + 1)
                        nv = str(next_cell.value or '').strip().lower()
                        if nv == 'sortie':
                            sim_row = r

        if not flt_row and not sim_row:
            return []

        # Step 3: Extract per-pilot data
        for pc in pilot_cols:
            name = str(pc.value).strip()
            col_start = pc.column  # Plan column for this pilot

            def get_val(row_num, offset):
                if row_num is None:
                    return 0
                cell = ws.cell(row=row_num, column=col_start + offset)
                return self._safe_int(cell.value)

            flt_plan = get_val(flt_row, 0)
            flt_done = get_val(flt_row, 1)
            flt_remain = get_val(flt_row, 2)
            sim_plan = get_val(sim_row, 0)
            sim_done = get_val(sim_row, 1)
            sim_remain = get_val(sim_row, 2)

            parsed.append({
                'name': name,
                'flt_plan': flt_plan, 'flt_done': flt_done, 'flt_remain': flt_remain,
                'sim_plan': sim_plan, 'sim_done': sim_done, 'sim_remain': sim_remain,
                'remarks': '',
            })

        return parsed

    def _parse_vertical_format(self, ws):
        """Fallback: parse vertical format with Pilot/Name column header."""
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
            for cell in row:
                val = str(cell.value or '').strip().lower()
                if val in ('pilot', 'name', 'pilot name', '조종사', '이름'):
                    header_row = row_idx
                    break
            if header_row:
                break
        if not header_row:
            return []

        col_map = {}
        for cell in ws[header_row]:
            val = str(cell.value or '').strip().lower()
            col = cell.column
            if val in ('pilot', 'name', 'pilot name', '조종사', '이름'):
                col_map['name'] = col
            elif val in ('flt plan', 'flight plan', 'flt_plan', 'flight_plan'):
                col_map['flt_plan'] = col
            elif val in ('flt done', 'flight done', 'flt_done', 'flight_done'):
                col_map['flt_done'] = col
            elif val in ('flt remain', 'flight remain', 'flt_remain', 'flight_remain'):
                col_map['flt_remain'] = col
            elif val in ('sim plan', 'sim_plan', 'simulator plan'):
                col_map['sim_plan'] = col
            elif val in ('sim done', 'sim_done', 'simulator done'):
                col_map['sim_done'] = col
            elif val in ('sim remain', 'sim_remain', 'simulator remain'):
                col_map['sim_remain'] = col
            elif val in ('plan',) and 'flt_plan' not in col_map:
                col_map['flt_plan'] = col
            elif val in ('done',) and 'flt_done' not in col_map:
                col_map['flt_done'] = col
            elif val in ('remain',) and 'flt_remain' not in col_map:
                col_map['flt_remain'] = col
            elif val in ('remarks', 'note', 'notes', '비고'):
                col_map['remarks'] = col

        if 'name' not in col_map:
            return []

        parsed = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            name_val = row[col_map['name'] - 1].value
            if not name_val or str(name_val).strip() == '':
                continue
            name_str = str(name_val).strip()
            if name_str.lower() in ('total', 'sum', '합계', '소계'):
                continue

            def safe_col(col_key):
                if col_key not in col_map:
                    return 0
                return self._safe_int(row[col_map[col_key] - 1].value)

            parsed.append({
                'name': name_str,
                'flt_plan': safe_col('flt_plan'), 'flt_done': safe_col('flt_done'),
                'flt_remain': safe_col('flt_remain'), 'sim_plan': safe_col('sim_plan'),
                'sim_done': safe_col('sim_done'), 'sim_remain': safe_col('sim_remain'),
                'remarks': str(row[col_map['remarks'] - 1].value or '') if 'remarks' in col_map else '',
            })
        return parsed

    def _parse_individual_status_sheet(self, ws):
        """Parse 'Individual Status' sheet to extract per-pilot, per-course completion dates/times.
        Returns list of dicts: {course_no, subject, category, pilot_name, completed_date, completed_time, contents}
        """
        results = []

        # Step 1: Find pilot names in row 5 (cols 8, 11, 14, 17, 20, 23 — 3-col stride)
        pilot_cols = []  # list of (col_index, pilot_name)
        for c in range(5, min(ws.max_column + 1, 40)):
            v = ws.cell(row=5, column=c).value
            if v and isinstance(v, str) and len(v.strip()) > 3 and not any(
                kw in v.lower() for kw in ('course', 'transition', 'status', 'rmaf', 'pilot', 'training',
                                             'contents', 'date', 'time', 'sim', 'flight', 'no.')
            ):
                # Clean newlines from multi-line cell names
                clean_name = ' '.join(v.strip().split())
                pilot_cols.append((c, clean_name))

        if len(pilot_cols) < 2:
            return []

        # Step 2: Scan data rows — identify course rows by Course No pattern (C-XX) in col 2
        for r in range(8, ws.max_row + 1):
            course_no_val = ws.cell(row=r, column=2).value
            if not course_no_val:
                continue
            course_no_str = str(course_no_val).strip()
            # Must be C-XX pattern
            if not course_no_str.upper().startswith('C-'):
                continue

            sim_no = ws.cell(row=r, column=3).value
            sim_subject = ws.cell(row=r, column=4).value
            flt_no = ws.cell(row=r, column=5).value
            flt_subject = ws.cell(row=r, column=6).value
            contents = str(ws.cell(row=r, column=7).value or '').strip()
            # Skip ditto marks
            if contents in ('\u3003', '"', '〃'):
                contents = ''

            # Determine if this is a SIM or Flight row
            is_sim = sim_subject is not None and str(sim_subject).strip() != ''
            is_flt = flt_subject is not None and str(flt_subject).strip() != ''

            if is_sim:
                subject = str(sim_subject).strip()
                category = 'sim'
            elif is_flt:
                subject = str(flt_subject).strip()
                category = 'flight'
            else:
                # Skip Option rows or empty
                continue

            # Also skip "CPT" only subject entries
            if subject.upper() in ('CPT', ''):
                continue

            # Step 3: Extract per-pilot date/time
            for (pcol, pname) in pilot_cols:
                date_val = ws.cell(row=r, column=pcol).value
                if date_val is None:
                    continue

                # Parse date
                date_str = ''
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%Y-%m-%d')
                elif isinstance(date_val, str) and date_val.strip():
                    date_str = date_val.strip()[:10]

                if not date_str:
                    continue

                # Parse time — SIM: pcol+1, Flight: pcol+2
                if is_sim:
                    time_val = ws.cell(row=r, column=pcol + 1).value
                else:
                    time_val = ws.cell(row=r, column=pcol + 2).value

                time_str = '1:00'
                if time_val:
                    if isinstance(time_val, datetime.timedelta):
                        total_sec = int(time_val.total_seconds())
                        h = total_sec // 3600
                        m = (total_sec % 3600) // 60
                        time_str = f'{h}:{m:02d}'
                    elif hasattr(time_val, 'strftime'):
                        time_str = time_val.strftime('%H:%M')
                        # Remove leading zero hour
                        if time_str.startswith('0') and len(time_str) > 4:
                            time_str = time_str[1:]
                    elif isinstance(time_val, str) and ':' in time_val:
                        parts = time_val.strip().split(':')
                        h = parts[0].lstrip('0') or '0'
                        time_str = h + ':' + parts[1]

                results.append({
                    'course_no': course_no_str,
                    'subject': subject,
                    'category': category,
                    'contents': contents,
                    'pilot_name': pname,
                    'completed_date': date_str,
                    'completed_time': time_str,
                })

        return results

    def _sync_individual_status(self, conn, individual_data, pilots):
        """Sync parsed Individual Status data to pilot_courses + pilot_training tables.
        - Auto-creates missing pilot_courses
        - Upserts pilot_training records
        Returns (synced_count, created_courses_count)
        """
        if not individual_data:
            return 0, 0

        # Build pilot lookup (name -> id) same fuzzy logic as weekly report
        pilot_lookup = {}
        for row in individual_data:
            pname = row['pilot_name']
            if pname in pilot_lookup:
                continue
            pname_lower = pname.lower()
            pilot_id = None
            # Pass 1: exact match
            for p in pilots:
                sn = (p.get('short_name') or '').lower()
                fn = (p.get('name') or '').lower()
                if (sn and sn == pname_lower) or (fn and fn == pname_lower):
                    pilot_id = p['id']
                    break
            # Pass 2: contains match
            if pilot_id is None:
                for p in pilots:
                    sn = (p.get('short_name') or '').lower()
                    fn = (p.get('name') or '').lower()
                    if (sn and (sn in pname_lower or pname_lower in sn)) or \
                       (fn and (fn in pname_lower or pname_lower in fn)):
                        pilot_id = p['id']
                        break
            pilot_lookup[pname] = pilot_id

        # Load existing courses
        existing_courses = dicts_from_rows(conn.execute(
            "SELECT * FROM pilot_courses ORDER BY sort_order, id"
        ).fetchall())

        # Build course lookup: subject (lowercase) -> course_id
        course_map = {}
        for c in existing_courses:
            course_map[c['subject'].lower()] = c['id']

        # Collect unique courses from data that may need to be created
        unique_courses = {}  # subject -> {course_no, category, contents}
        for row in individual_data:
            subj = row['subject']
            if subj.lower() not in course_map and subj not in unique_courses:
                unique_courses[subj] = {
                    'course_no': row['course_no'],
                    'category': row['category'],
                    'contents': row['contents'],
                }

        # Auto-create missing courses
        created_count = 0
        max_sort = 0
        if existing_courses:
            max_sort = max(c.get('sort_order', 0) or 0 for c in existing_courses)

        for subj, info in unique_courses.items():
            max_sort += 1
            # Extract seq_no from subject (e.g., TR-1S -> 1, INST-2 -> 2)
            seq_no = 0
            m = re.search(r'(\d+)', subj)
            if m:
                seq_no = int(m.group(1))

            cur = conn.execute(
                """INSERT INTO pilot_courses (course_no, category, seq_no, subject, contents, duration, sort_order)
                   VALUES (?,?,?,?,?,?,?)""",
                (info['course_no'], info['category'], seq_no, subj,
                 info['contents'], '1:00', max_sort)
            )
            course_map[subj.lower()] = cur.lastrowid
            created_count += 1

        # Upsert pilot_training records
        synced_count = 0
        for row in individual_data:
            pilot_id = pilot_lookup.get(row['pilot_name'])
            if not pilot_id:
                continue
            course_id = course_map.get(row['subject'].lower())
            if not course_id:
                continue

            # Check existing
            existing = conn.execute(
                "SELECT id FROM pilot_training WHERE pilot_id=? AND course_id=?",
                (pilot_id, course_id)
            ).fetchone()

            if existing:
                conn.execute(
                    """UPDATE pilot_training SET completed_date=?, completed_time=?,
                       updated_at=CURRENT_TIMESTAMP WHERE pilot_id=? AND course_id=?""",
                    (row['completed_date'], row['completed_time'], pilot_id, course_id)
                )
            else:
                conn.execute(
                    """INSERT INTO pilot_training (pilot_id, course_id, completed_date, completed_time, notes)
                       VALUES (?,?,?,?,?)""",
                    (pilot_id, course_id, row['completed_date'], row['completed_time'], '')
                )
            synced_count += 1

        return synced_count, created_count

    @require_auth()
    def get(self):
        conn = get_db()
        uploads = dicts_from_rows(conn.execute(
            "SELECT id, filename, original_filename, uploaded_by, report_date, file_size, row_count, notes, created_at FROM weekly_uploads ORDER BY created_at DESC"
        ).fetchall())
        conn.close()
        self.success(uploads)

    @require_auth(roles=['admin', 'ojt_admin', 'instructor'])
    def post(self):
        files = self.request.files.get("file", [])
        if not files:
            return self.error("No file uploaded")

        f = files[0]
        orig_name = f["filename"]
        ext = os.path.splitext(orig_name)[1].lower()
        if ext not in ('.xlsx', '.xls', '.xlsm'):
            return self.error("Only .xlsx, .xls, .xlsm files are supported")
        if len(f["body"]) > 10 * 1024 * 1024:
            return self.error("File too large (max 10MB)")

        # Parse Excel
        try:
            import openpyxl
        except ImportError:
            return self.error("openpyxl not installed on server")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(f["body"]), data_only=True)
            parsed_rows = []

            # --- Try "Weekly Report" sheet first (RMAF horizontal format) ---
            ws = None
            for sn in wb.sheetnames:
                if 'weekly' in sn.lower() and 'report' in sn.lower():
                    ws = wb[sn]
                    break

            if ws:
                parsed_rows = self._parse_weekly_report_sheet(ws)

            # --- Fallback: try active sheet with vertical format ---
            if not parsed_rows:
                ws = wb.active
                parsed_rows = self._parse_vertical_format(ws)

            # --- Fallback: try all sheets ---
            if not parsed_rows:
                for sn in wb.sheetnames:
                    parsed_rows = self._parse_weekly_report_sheet(wb[sn])
                    if parsed_rows:
                        break
                    parsed_rows = self._parse_vertical_format(wb[sn])
                    if parsed_rows:
                        break

            wb.close()

        except Exception as ex:
            import traceback
            traceback.print_exc()
            return self.error(f"Failed to parse Excel file: {str(ex)}")

        if not parsed_rows:
            return self.error("No data rows found in Excel file")

        # Save file to disk (for local access) and keep binary for DB
        os.makedirs(WEEKLY_UPLOAD_DIR, exist_ok=True)
        fname = f"weekly_{uuid.uuid4().hex[:8]}{ext}"
        fpath = os.path.join(WEEKLY_UPLOAD_DIR, fname)
        file_binary = f["body"]
        with open(fpath, "wb") as fp:
            fp.write(file_binary)

        # Get uploader info
        user = get_current_user(self)
        uploader = user['name'] if user else 'Unknown'
        report_date = self.get_argument('report_date', datetime.date.today().isoformat())
        notes = self.get_argument('notes', '')

        # Save to DB (including file binary for persistence across redeploys)
        conn = get_db()
        try:
            from database import IS_POSTGRES
            if IS_POSTGRES:
                import psycopg2
                cur = conn.execute(
                    """INSERT INTO weekly_uploads
                       (filename, original_filename, uploaded_by, report_date, file_size, row_count, notes, file_data)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (fname, orig_name, uploader, report_date, len(file_binary), len(parsed_rows), notes,
                     psycopg2.Binary(file_binary))
                )
            else:
                cur = conn.execute(
                    """INSERT INTO weekly_uploads
                       (filename, original_filename, uploaded_by, report_date, file_size, row_count, notes, file_data)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (fname, orig_name, uploader, report_date, len(file_binary), len(parsed_rows), notes,
                     file_binary)
                )
            upload_id = cur.lastrowid

            # Match pilot names and save parsed data
            pilots = dicts_from_rows(conn.execute(
                "SELECT id, name, short_name FROM pilots WHERE status='active'"
            ).fetchall())

            for pr in parsed_rows:
                # Try to match pilot by short_name, name (exact then contains)
                pilot_id = None
                pr_lower = pr['name'].lower()
                # Pass 1: exact match
                for p in pilots:
                    sn = (p.get('short_name') or '').lower()
                    fn = (p.get('name') or '').lower()
                    if (sn and sn == pr_lower) or (fn and fn == pr_lower):
                        pilot_id = p['id']
                        break
                # Pass 2: contains match (excel name in DB name or DB name in excel name)
                if pilot_id is None:
                    for p in pilots:
                        sn = (p.get('short_name') or '').lower()
                        fn = (p.get('name') or '').lower()
                        if (sn and (sn in pr_lower or pr_lower in sn)) or \
                           (fn and (fn in pr_lower or pr_lower in fn)):
                            pilot_id = p['id']
                            break
                conn.execute(
                    """INSERT INTO weekly_report_data
                       (upload_id, pilot_id, pilot_name, flt_plan, flt_done, flt_remain,
                        sim_plan, sim_done, sim_remain, remarks)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (upload_id, pilot_id, pr['name'],
                     pr['flt_plan'], pr['flt_done'], pr['flt_remain'],
                     pr['sim_plan'], pr['sim_done'], pr['sim_remain'],
                     pr['remarks'])
                )

            conn.commit()
            upload = dict_from_row(conn.execute(
                "SELECT id, filename, original_filename, uploaded_by, report_date, file_size, row_count, notes, created_at FROM weekly_uploads WHERE id=?",
                (upload_id,)
            ).fetchone())
            matched_names = []
            unmatched_names = []
            for r in parsed_rows:
                r_lower = r['name'].lower().strip()
                found = False
                matched_to = None
                for p in pilots:
                    sn = (p.get('short_name') or '').lower().strip()
                    fn = (p.get('name') or '').lower().strip()
                    if (sn and sn == r_lower) or (fn and fn == r_lower):
                        found = True
                        matched_to = p.get('name') or p.get('short_name')
                        break
                if not found:
                    # Try contains match (same logic as insertion above)
                    for p in pilots:
                        sn = (p.get('short_name') or '').lower().strip()
                        fn = (p.get('name') or '').lower().strip()
                        if (sn and (sn in r_lower or r_lower in sn)) or \
                           (fn and (fn in r_lower or r_lower in fn)):
                            found = True
                            matched_to = p.get('name') or p.get('short_name')
                            break
                if found:
                    matched_names.append({'excel_name': r['name'], 'db_name': matched_to})
                else:
                    # Collect DB pilot names for suggestion
                    all_db_names = [p.get('name') or p.get('short_name') or '' for p in pilots]
                    unmatched_names.append({'excel_name': r['name'], 'db_pilots': all_db_names[:10]})
            matched = len(matched_names)

            # --- Parse Individual Status sheet and sync to pilot_training ---
            individual_synced = 0
            courses_created = 0
            ind_error = None
            ind_debug = {}
            try:
                wb2 = openpyxl.load_workbook(io.BytesIO(file_binary), data_only=True)
                ind_debug['sheets'] = wb2.sheetnames
                ind_ws = None
                for sn in wb2.sheetnames:
                    if 'individual' in sn.lower() and 'status' in sn.lower():
                        ind_ws = wb2[sn]
                        ind_debug['found_sheet'] = sn
                        break
                if ind_ws:
                    individual_data = self._parse_individual_status_sheet(ind_ws)
                    ind_debug['parsed_count'] = len(individual_data)
                    ind_debug['pilot_names'] = list(set(r['pilot_name'] for r in individual_data))[:6] if individual_data else []
                    if individual_data:
                        individual_synced, courses_created = self._sync_individual_status(conn, individual_data, pilots)
                        conn.commit()
                        ind_debug['synced'] = individual_synced
                        ind_debug['created'] = courses_created
                else:
                    ind_debug['error'] = 'No Individual Status sheet found'
                wb2.close()
            except Exception as ex2:
                import traceback
                traceback.print_exc()
                ind_error = str(ex2)
                ind_debug['exception'] = ind_error
                print(f"[WARN] Individual Status parsing failed: {ex2}")

            msg = f"Uploaded successfully: {len(parsed_rows)} rows parsed"
            if individual_synced > 0:
                msg += f", {individual_synced} training records synced"
            if courses_created > 0:
                msg += f", {courses_created} new courses created"
            if ind_error:
                msg += f" (Individual Status error: {ind_error})"
            self.success({'upload': upload, 'parsed_rows': parsed_rows, 'matched': matched,
                          'unmatched_names': unmatched_names, 'matched_names': matched_names,
                          'individual_synced': individual_synced, 'courses_created': courses_created,
                          'ind_debug': ind_debug},
                         msg)
        except Exception as e:
            conn.rollback()
            import traceback
            traceback.print_exc()
            self.error(f"Failed to save upload data: {str(e)}", 500)
        finally:
            conn.close()


class WeeklyUploadDetailHandler(BaseHandler):
    """GET one upload detail, DELETE remove upload"""

    @require_auth()
    def get(self, upload_id):
        conn = get_db()
        upload = dict_from_row(conn.execute(
            "SELECT id, filename, original_filename, uploaded_by, report_date, file_size, row_count, notes, created_at FROM weekly_uploads WHERE id=?",
            (upload_id,)
        ).fetchone())
        if not upload:
            conn.close()
            return self.error("Upload not found", 404)
        data = dicts_from_rows(conn.execute(
            "SELECT * FROM weekly_report_data WHERE upload_id=? ORDER BY id", (upload_id,)
        ).fetchall())
        conn.close()
        self.success({'upload': upload, 'data': data})

    @require_auth(roles=['admin', 'ojt_admin'])
    def delete(self, upload_id):
        conn = get_db()
        upload = dict_from_row(conn.execute("SELECT id, filename FROM weekly_uploads WHERE id=?", (upload_id,)).fetchone())
        if not upload:
            conn.close()
            return self.error("Upload not found", 404)

        # Delete file
        fpath = os.path.join(WEEKLY_UPLOAD_DIR, upload['filename'])
        if os.path.exists(fpath):
            os.remove(fpath)

        # Delete DB records
        conn.execute("DELETE FROM weekly_report_data WHERE upload_id=?", (upload_id,))
        conn.execute("DELETE FROM weekly_uploads WHERE id=?", (upload_id,))
        conn.commit()
        conn.close()
        self.success(message="Upload deleted")


class WeeklyUploadDownloadHandler(BaseHandler):
    """GET download original Excel file (supports token in query param for window.open)"""

    def get(self, upload_id):
        # Support token in query param for file downloads via window.open
        from auth import decode_token
        token = self.get_argument('token', None)
        auth = self.request.headers.get("Authorization", "")
        user = None
        if auth.startswith("Bearer "):
            user = decode_token(auth[7:])
        elif token:
            user = decode_token(token)
        if not user:
            self.set_status(401)
            self.write({"error": "Authentication required"})
            return

        conn = get_db()
        upload = dict_from_row(conn.execute(
            "SELECT id, filename, original_filename FROM weekly_uploads WHERE id=?", (upload_id,)
        ).fetchone())
        conn.close()
        if not upload:
            return self.error("Upload not found", 404)

        fpath = os.path.join(WEEKLY_UPLOAD_DIR, upload['filename'])
        file_data = None

        # Try filesystem first
        if os.path.exists(fpath):
            with open(fpath, 'rb') as fp:
                file_data = fp.read()
        else:
            # Fallback: load from DB (survives redeploys)
            conn2 = get_db()
            row = conn2.execute("SELECT file_data FROM weekly_uploads WHERE id=?", (upload_id,)).fetchone()
            conn2.close()
            if row and row['file_data']:
                raw = row['file_data']
                file_data = bytes(raw) if isinstance(raw, (memoryview, bytearray)) else raw

        if not file_data:
            return self.error("File not found on server", 404)

        orig = upload['original_filename']
        ext = orig.rsplit('.', 1)[-1].lower() if '.' in orig else 'xlsx'
        ct = 'application/vnd.ms-excel.sheet.macroEnabled.12' if ext == 'xlsm' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        self.set_header('Content-Type', ct)
        self.set_header('Content-Disposition', f'attachment; filename="{orig}"')
        self.write(file_data)
        self.finish()


class WeeklyUploadLatestHandler(BaseHandler):
    """GET latest weekly report data (from most recent upload)"""

    @require_auth()
    def get(self):
        conn = get_db()
        latest = conn.execute(
            "SELECT id FROM weekly_uploads ORDER BY created_at DESC LIMIT 1"
        ).fetchone()
        if not latest:
            conn.close()
            return self.success([])

        data = dicts_from_rows(conn.execute(
            """SELECT wrd.*, wu.report_date, wu.original_filename
               FROM weekly_report_data wrd
               JOIN weekly_uploads wu ON wrd.upload_id = wu.id
               WHERE wrd.upload_id=? ORDER BY wrd.id""",
            (latest['id'],)
        ).fetchall())
        conn.close()
        self.success(data)
